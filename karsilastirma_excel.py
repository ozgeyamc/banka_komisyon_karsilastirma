"""
Bankalardan çekilen verileri karşılaştırmalı Excel formatında yazan modül.
Eşleştirme: işlem tipi + tutar aralığı standardizasyonu ile yapılır.
"""

import os
import re
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Tuple, Optional
from models import UcretSatiri

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_DOSYA_ADI = "komisyon_karsilastirma.xlsx"
SHEET_ADI = "karşılaştırma"

BANKALAR = ["GARANTİ", "İŞBANKASI", "AKBANK", "YAPIKREDI"]
BANKA_TAMAD = {
    "GARANTİ":   "Garanti BBVA",
    "İŞBANKASI": "İş Bankası",
    "AKBANK":    "Akbank",
    "YAPIKREDI": "Yapı ve Kredi Bankası",
}

# Başlık (satır 1) renkleri
BANKA_RENKLER = {
    "GARANTİ":   {"bg": "00B050", "fg": "FFFFFF"},
    "İŞBANKASI": {"bg": "012169", "fg": "FFFFFF"},
    "AKBANK":    {"bg": "FF0000", "fg": "FFFFFF"},
    "YAPIKREDI": {"bg": "003087", "fg": "FFD700"},
}

# Alt başlık (Mobil/Şube - satır 2) renkleri - görseldeki pastel tonlar
BANKA_ALT_RENKLER = {
    "GARANTİ":   {"bg": "C6E0B4", "fg": "FFFFFF"},
    "İŞBANKASI": {"bg": "BDD7EE", "fg": "FFFFFF"},
    "AKBANK":    {"bg": "F4B183", "fg": "FFFFFF"},
    "YAPIKREDI": {"bg": "8EA9C1", "fg": "FFFFFF"},
}

# Hücre veri metni rengi - bankaya göre (görseldeki gibi)
BANKA_VERI_RENK = {
    "GARANTİ":   "00B050",
    "İŞBANKASI": "0070C0",
    "AKBANK":    "FF0000",
    "YAPIKREDI": "1F3864",
}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NOT_KOLONU = 12  # L sütunu - notlar buraya yazılır


def _bugun() -> str:
    tz = timezone(timedelta(hours=3))
    return datetime.now(tz).strftime("%d.%m.%Y %H:%M")


def _deger(s: UcretSatiri) -> str:
    t = (s.asgari_tutar or "").strip()
    o = (s.asgari_oran  or "").strip()
    parts = []
    if t: parts.append(t)
    if o: parts.append(o)
    return " / ".join(parts) if parts else ""


def _norm(m: str) -> str:
    m = m.lower().strip()
    for a, b in [("ı","i"),("ğ","g"),("ü","u"),("ş","s"),("ö","o"),("ç","c"),
                 ("â","a"),("î","i"),("û","u"),("i̇","i")]:
        m = m.replace(a, b)
    return re.sub(r"\s+", " ", m).strip()


def _extract_tutar(m: str) -> Optional[str]:
    """
    Metindeki tutar aralığını standart bir etikete çevirir.

    Eski yaklaşım "8300-399000" gibi bir alt-string arıyordu. Ancak Akbank
    ("8.300,01 TL - 399.000 TL arasında") ve Yapı Kredi ("8.300,01 TL – 399.000 TL")
    sayılar arasına "TL" gibi metin soktuğu için bu substring hiç oluşmuyordu ve
    tutar tespiti başarısız oluyordu. Ayrıca Akbank'ın "ve Altı" / "ve Üstü" gibi
    kelime bazlı sınır ifadeleri de hiç yakalanmıyordu.

    Yeni yaklaşım: metindeki sayıları çıkarıp büyüklüklerine göre, ve
    "ve altı"/"ve üstü" gibi anahtar kelimelere göre bandı belirler.
    """
    t = m.replace("–", "-").replace("—", "-").lower()
    tn = _norm(m)  # türkçe karakterleri sadeleştirilmiş hali (anahtar kelime kontrolü için)

    # Kasa boyutu / gram aralığı etiketleri (öncelikli, sayısal değil)
    for pat, label in [
        (r"buyuk|büyük", "büyük"),
        (r"orta(?!k)", "orta"),
        (r"kucuk|küçük", "küçük"),
        (r"1[-\s]*10\s*g", "1-10gr"),
        (r"11[-\s]*100\s*g", "11-100gr"),
    ]:
        if re.search(pat, t, re.IGNORECASE):
            return label

    # Metindeki tüm sayıları normalize ederek çıkar (binlik ayracı sil, ondalık at)
    nums = []
    for raw in re.findall(r"\d{1,3}(?:\.\d{3})*(?:,\d+)?", t):
        cleaned = re.sub(r"\.(?=\d{3}(\D|$))", "", raw)
        cleaned = re.sub(r",\d+$", "", cleaned)
        if cleaned.isdigit():
            nums.append(int(cleaned))
    if not nums:
        return None

    ust_ifade = any(k in tn for k in ["ve ustu", "ve uzeri", "uzeri", "ustu"])
    alt_ifade = any(k in tn for k in ["ve alti", "altinda", "ve altinda"])

    if ust_ifade:
        return "399000+"
    if alt_ifade:
        return "0-8300"

    if len(nums) >= 2:
        ilk = nums[0]
        if ilk <= 0:
            return "0-8300"
        if ilk >= 300000:
            return "399000+"
        return "8300-399000"

    ilk = nums[0]
    if ilk <= 0:
        return "0-8300"
    if ilk >= 300000:
        return "399000+"
    if ilk <= 8300:
        return "0-8300"
    return "8300-399000"


def _norm_deger(d: str) -> str:
    if not d:
        return ""
    d = d.strip().upper()
    d = re.sub(r"(\d),(\d)", r"\1.\2", d)
    d = d.replace(" ", "").replace("TL", "TRY")
    d = re.sub(r"[^0-9A-Z%./]", "", d)
    try:
        num = re.search(r"[\d.]+", d)
        if num:
            val = float(num.group())
            d = d[:num.start()] + f"{val:.2f}" + d[num.end():]
    except Exception:
        pass
    return d


def _standart_anahtar(masraf: str) -> Tuple[str, str, str]:
    # ... (değişmedi, önceki koddaki mantık aynen korunur)
    yk_kat = ""
    if " | " in masraf:
        parts = masraf.split(" | ", 1)
        yk_kat = _norm(parts[0])
        masraf = parts[1]

    n = _norm(masraf)
    tutar = _extract_tutar(masraf)

    def kanal():
        if any(k in n for k in ["mobil","internet","iscep","online","dijital"]):
            return "mobil"
        if any(k in n for k in ["sube","şube","gise","gişe","cozum","çözüm","iletisim"]):
            return "sube"
        if "atm" in n:
            return "mobil"
        return ""

    if yk_kat:
        if "eft" in yk_kat or "fast" in yk_kat:
            if "fast" in yk_kat:
                label = f"FAST - {tutar} TRY" if tutar else "FAST"
                return "FAST", label, kanal()
            label = f"EFT - {tutar} TRY" if tutar else "EFT Gönderimi"
            return "EFT Gönderimi", label, kanal()
        if "havale" in yk_kat and "atm" not in yk_kat:
            if "duzenli" in n or "talimat" in n:
                label = f"Düzenli Havale - {tutar} TRY" if tutar else "Düzenli Havale"
                return "Havale Gönderimi", label, kanal()
            label = f"Havale - {tutar} TRY" if tutar else "Havale Gönderimi"
            return "Havale Gönderimi", label, kanal()
        if "atm" in yk_kat and "havale" in yk_kat:
            label = f"Havale - {tutar} TRY" if tutar else "Havale Gönderimi"
            return "Havale Gönderimi", label, "mobil"
        if "kiralik kasa" in yk_kat or "kiralık kasa" in yk_kat:
            if "depozito" in n:
                return "Kiralık Kasa", f"Kasa Depozito - {tutar or n}", ""
            return "Kiralık Kasa", f"Yıllık Kasa Ücreti - {tutar or n}", ""
        if "senet" in yk_kat:
            if "iade" in n:
                return "Senet İade Ücreti", "Senet İade Ücreti", ""
            if "protesto" in n and "kaldir" in n:
                return "Senet Protesto İşlemleri Ücreti", "Senet Protesto Kaldırma -", ""
            if "protesto" in n:
                return "Senet Protesto İşlemleri Ücreti", "Senet Protesto -", ""
            if "bankamizda" in n or "bankamızda" in n:
                return "Senet Tahsile Alma Ücreti", "Aynı Banka Senet Tahsili -", ""
            return "Senet Tahsile Alma Ücreti", "Muhabir Banka Senet Tahsili -", ""
        if "cek" in yk_kat or "çek" in yk_kat:
            if "ayni sube" in n or "aynı şube" in n:
                return "Çek Tahsilat Ücreti", "Aynı Banka Çeki -", ""
            if "baska sube" in n or "başka şube" in n:
                return "Çek Tahsilat Ücreti", "Diğer Banka Çeki -", ""
            if "karsiliksiz" in n:
                return "Çek Belgelendirme ve Düzeltme Ücreti", "Karşılıksız Çek Belgelendirme -", ""
            if "iade" in n:
                return "Çek İade Ücreti", "Çek İade Ücreti", ""
            if "yaprak" in n or "defteri" in n:
                return "Çek Defteri ve Çek Düzenleme Ücreti", "Çek Defteri (Yaprak Başı) -", ""
            if "duzenleme" in n:
                return "Çek Defteri ve Çek Düzenleme Ücreti", "Çek Düzenleme -", ""
            return "Çek Tahsilat Ücreti", "Çek Tahsilat", ""
        if "ortak atm" in yk_kat:
            if "bakiye" in n:
                return "Bakiye Sorma - Yurtiçi - ATM", "Bakiye Sorma - Yurtiçi - ATM", ""
        if "kurum tahsilat" in yk_kat or "kredi karti islem" in yk_kat:
            if "fatura" in n:
                return "Fatura Ödemeleri - Kart", "Fatura Ödemeleri", kanal()
            if "sgk" in n or "prim" in n:
                return "SGK Prim Ödemeleri", "SGK Prim Ödemeleri", kanal()
        if "findeks" in yk_kat or ("kredi" in yk_kat and "risk" in yk_kat):
            return ("Üçüncü Kişi ve Kuruluşlardan Temin Edilecek Rapor Ücretleri - Kredi Risk Raporu",
                    "Üçüncü Kişi ve Kuruluşlardan Temin Edilecek Rapor Ücretleri - Kredi Risk Raporu",
                    kanal())
        if "referans" in yk_kat:
            return "Mevduat Araştırma", "Referans Mektubu -", kanal()
        return None, None, ""

    if "eft" in n and "swift" not in n and "uluslararasi" not in n:
        if "duzenli" in n or "supurme" in n:
            label = f"Düzenli EFT - {tutar} TRY" if tutar else "Düzenli EFT"
            return "EFT Gönderimi", label, kanal()
        if "odenmesi" in n or "isme gelen" in n:
            return "EFT Gönderimi", "EFT İsme Gelen", "sube"
        label = f"EFT - {tutar} TRY" if tutar else "EFT Gönderimi"
        return "EFT Gönderimi", label, kanal()

    if "havale" in n and "uluslararasi" not in n and "swift" not in n:
        if "duzenli" in n or "supurme" in n:
            label = f"Düzenli Havale - {tutar} TRY" if tutar else "Düzenli Havale"
            return "Havale Gönderimi", label, kanal()
        if "cebe para" in n:
            return "Havale Gönderimi", "Cebe Para Gönderme", kanal()
        label = f"Havale - {tutar} TRY" if tutar else "Havale Gönderimi"
        return "Havale Gönderimi", label, kanal()

    if "fast" in n:
        label = f"FAST - {tutar} TRY" if tutar else "FAST"
        return "FAST", label, kanal()

    if "kasa" in n:
        if "depozito" in n:
            return "Kiralık Kasa", f"Kasa Depozito - {tutar or 'genel'}", ""
        return "Kiralık Kasa", f"Yıllık Kasa Ücreti - {tutar or 'genel'}", ""

    if "altin" in n or "kiymetli maden" in n or "ats" in n:
        label = f"Altın Transfer - {tutar}" if tutar else "Altın Transfer"
        return "Kıymetli Maden", label, kanal()

    if "hgs" in n:
        if "etiket" in n:
            return "HGS Etiket Bedeli", "HGS Etiket Bedeli", ""
        return "HGS", "HGS Kart Ücreti", ""

    if "sans oyunu" in n or "piyango" in n:
        return "Şans Oyunu Ödemeleri Aracılık", "Şans Oyunu Ödemeleri Aracılık", kanal()

    if "fatura" in n and "kredi" not in n and "ekstre" not in n:
        return "Fatura Ödemeleri - Kart", "Fatura Ödemeleri", kanal()

    if "sgk" in n:
        label = f"SGK - {tutar} TRY" if tutar else "SGK Prim Ödemeleri"
        return "SGK Prim Ödemeleri", label, kanal()

    if "vergi" in n and "kredi" not in n:
        label = f"Vergi - {tutar} TRY" if tutar else "Vergi Tahsilat"
        return "Vergi Tahsilat Aracılık", label, kanal()

    if "aidat" in n:
        label = f"Aidat - {tutar} TRY" if tutar else "Aidat Ödemeleri"
        return "Aidat Ödemeleri Aracılık", label, kanal()

    if "okul" in n:
        label = f"Özel Okul - {tutar} TRY" if tutar else "Özel Okul Ödeme"
        return "Özel Okul Ödeme", label, kanal()

    if "telefon" in n:
        label = f"Telefon - {tutar} TRY" if tutar else "Telefon Ödemeleri"
        return "Telefon Ödemeleri Aracılık", label, kanal()

    if "arsiv" in n:
        return "Arşiv Araştırma Ücreti", "Arşiv Araştırma Ücreti", kanal()

    if "referans" in n or "itibar" in n or "niyet" in n:
        return "Mevduat Araştırma", "Referans Mektubu -", kanal()
    if "hesap ozeti" in n:
        return "Mevduat Araştırma", "Hesap Özeti Verilmesi -", kanal()
    if "hesap arastirma" in n:
        return "Mevduat Araştırma", "Hesap Araştırma Talebi -", kanal()
    if "borcu yok" in n:
        return "Mevduat Araştırma", "Borcu Yoktur Yazısı", kanal()
    if "vize" in n and "okul" in n:
        return "Mevduat Araştırma", "Vize ve Özel Okullar için Düzenlenen Mektuplar -", kanal()

    if "bakiye" in n and "atm" in n:
        if "yurtici" in n or ("yurt" in n and "dis" not in n):
            return "Bakiye Sorma - Yurtiçi - ATM", "Bakiye Sorma - Yurtiçi - ATM", ""
        return "Bakiye Sorma - Yurtdışı - ATM", "Bakiye Sorma - Yurtdışı - ATM", ""

    if "kkb" in n or ("kredi" in n and "risk" in n) or ("ucuncu" in n and "rapor" in n):
        return ("Üçüncü Kişi ve Kuruluşlardan Temin Edilecek Rapor Ücretleri - Kredi Risk Raporu",
                "Üçüncü Kişi ve Kuruluşlardan Temin Edilecek Rapor Ücretleri - Kredi Risk Raporu",
                kanal())

    if ("cek" in n or "çek" in n) and ("defteri" in n or "yaprak" in n or "teslim" in n):
        return "Çek Defteri ve Çek Düzenleme Ücreti", "Çek Defteri (Yaprak Başı) -", ""
    if ("cek" in n or "çek" in n) and "duzenleme" in n:
        if "ozel" in n or "nitelik" in n:
            return "Çek Defteri ve Çek Düzenleme Ücreti", "Özel Nitelikli Çek Düzenleme -", ""
        return "Çek Defteri ve Çek Düzenleme Ücreti", "Çek Düzenleme -", ""

    if ("cek" in n or "çek" in n) and "iade" in n:
        return "Çek İade Ücreti", "Çek İade Ücreti", ""

    if ("cek" in n or "çek" in n) and ("tahsil" in n or "odeme" in n):
        if "ayni banka" in n or "aynı banka" in n:
            return "Çek Tahsilat Ücreti", "Aynı Banka Çeki -", ""
        if "diger banka" in n or "baska banka" in n:
            return "Çek Tahsilat Ücreti", "Diğer Banka Çeki -", ""
        if "doviz" in n or "yp" in n:
            return "Çek Tahsilat Ücreti", "Döviz Çekleri Tahsilatı (Diğer Banka) -", ""
        return "Çek Tahsilat Ücreti", "Çek Tahsilat", ""

    if ("cek" in n or "çek" in n) and ("belgelend" in n or "karsiliksiz" in n or "duzeltme" in n):
        if "karsiliksiz" in n:
            return "Çek Belgelendirme ve Düzeltme Ücreti", "Karşılıksız Çek Belgelendirme -", ""
        return "Çek Belgelendirme ve Düzeltme Ücreti", "Çek Düzeltme Hakkı -", ""

    if "senet" in n and "iade" in n:
        return "Senet İade Ücreti", "Senet İade Ücreti", ""

    if "senet" in n and "protesto" in n:
        if "kaldir" in n:
            return "Senet Protesto İşlemleri Ücreti", "Senet Protesto Kaldırma -", ""
        return "Senet Protesto İşlemleri Ücreti", "Senet Protesto -", ""

    if "senet" in n and ("tahsil" in n or "tahsile" in n):
        if "ayni" in n:
            return "Senet Tahsile Alma Ücreti", "Aynı Banka Senet Tahsili -", ""
        return "Senet Tahsile Alma Ücreti", "Muhabir Banka Senet Tahsili -", ""

    return None, None, ""


KATEGORI_SIRA = [
    "EFT Gönderimi",
    "Havale Gönderimi",
    "FAST",
    "Kiralık Kasa",
    "Kıymetli Maden",
    "Üçüncü Kişi ve Kuruluşlardan Temin Edilecek Rapor Ücretleri - Kredi Risk Raporu",
    "Fatura Ödemeleri - Kart",
    "SGK Prim Ödemeleri",
    "HGS Etiket Bedeli",
    "Şans Oyunu Ödemeleri Aracılık",
    "Aidat Ödemeleri Aracılık",
    "Özel Okul Ödeme",
    "Telefon Ödemeleri Aracılık",
    "Vergi Tahsilat Aracılık",
    "Arşiv Araştırma Ücreti",
    "Mevduat Araştırma",
    "Bakiye Sorma - Yurtiçi - ATM",
    "Bakiye Sorma - Yurtdışı - ATM",
    "Çek Defteri ve Çek Düzenleme Ücreti",
    "Çek İade Ücreti",
    "Çek Tahsilat Ücreti",
    "Çek Belgelendirme ve Düzeltme Ücreti",
    "Senet İade Ücreti",
    "Senet Protesto İşlemleri Ücreti",
    "Senet Tahsile Alma Ücreti",
]

# Görselde sağ tarafta görülen kırmızı italik notlar. Kategoriye göre eşlenir.
KATEGORI_NOTLARI: Dict[str, List[str]] = {
    "EFT Gönderimi": [
        "Para çekmeyi, para yatırmayı eklemedim",
        "Akbank'ta bireysel kotalar var havale/EFT için",
        "Kredi Risk Raporumuz zamanlanabilir",
        "Şubeden kıymetli maden teslimi %10'a kadar",
    ],
    "Fatura Ödemeleri - Kart": [
        "Fatura ödemeleri üst tier için %3,5'a çıkabiliyor",
        "Borcu Yoktur Yazısı almak için ücret alınabiliyor",
        "Mevduat Araştırma Ücretleri zamlanabilir",
        "Güvenli Araç Alım Satım YKB 97,53 TRY",
    ],
}


def karsilastirma_excel_yaz(
    banka_verileri: Dict[str, List[UcretSatiri]],
    dosya_yolu: str = EXCEL_DOSYA_ADI,
) -> None:
    if os.path.exists(dosya_yolu):
        os.remove(dosya_yolu)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_ADI
    MAX_COL = 9  # A..I (A=masraf, B-I=4 banka x 2 kanal)

    def sb(cell):
        cell.border = BORDER

    # ---- A1 boş üst-sol köşe ----
    ws.merge_cells("A1:A2")
    c = ws["A1"]
    c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    c.font = Font(color="FFFFFF", bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    sb(c)

    # ---- Banka başlıkları (satır 1) + Mobil/Şube alt başlıkları (satır 2, italik) ----
    col = 2
    for banka in BANKALAR:
        r = BANKA_RENKLER[banka]
        ar = BANKA_ALT_RENKLER[banka]
        fill = PatternFill(start_color=r["bg"], end_color=r["bg"], fill_type="solid")
        alt_fill = PatternFill(start_color=ar["bg"], end_color=ar["bg"], fill_type="solid")
        l1, l2 = get_column_letter(col), get_column_letter(col + 1)
        ws.merge_cells(f"{l1}1:{l2}1")
        c1 = ws[f"{l1}1"]
        c1.value = BANKA_TAMAD[banka]
        c1.fill = fill
        c1.font = Font(color=r["fg"], bold=True, size=11)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        sb(c1)
        for j, kanal_adi in enumerate(["Mobil", "Şube"]):
            c2 = ws.cell(row=2, column=col + j)
            c2.value = kanal_adi
            c2.fill = alt_fill
            c2.font = Font(color=ar["fg"], bold=True, italic=True, size=10)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            sb(c2)
        col += 2

    ws.column_dimensions["A"].width = 42
    for i in range(2, MAX_COL + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    # Notlar sütunu daha geniş
    ws.column_dimensions[get_column_letter(NOT_KOLONU)].width = 55
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    kat_satirlar: Dict[str, List[str]] = {}
    veri: Dict[Tuple[str, str], Dict[str, Dict[str, str]]] = {}

    for banka in BANKALAR:
        for s in banka_verileri.get(banka, []):
            kat, isim, kanal = _standart_anahtar(s.masraf)
            if kat is None:
                continue
            if kanal not in ("mobil", "sube"):
                k2 = (s.kanal or "").lower()
                kanal = k2 if k2 in ("mobil", "sube") else "mobil"

            key = (kat, isim)
            if kat not in kat_satirlar:
                kat_satirlar[kat] = []
            if isim not in kat_satirlar[kat]:
                kat_satirlar[kat].append(isim)

            if key not in veri:
                veri[key] = {}
            if banka not in veri[key]:
                veri[key][banka] = {"mobil": "", "sube": ""}

            d = _deger(s)
            depozito = getattr(s, "depozito", "") or ""
            if d:
                mevcut = veri[key][banka][kanal]
                if depozito:
                    d = f"{d}\nDepozito {depozito}"
                if mevcut and depozito and "Depozito" not in mevcut:
                    veri[key][banka][kanal] = d
                else:
                    veri[key][banka][kanal] = d if d else mevcut

    KAT_FONT = Font(bold=True, size=10)
    DATA_FONT = Font(size=10)
    MASRAF_FONT = Font(size=10)

    yazilan = set()
    row = 3
    toplam = 0
    kategori_baslangic_satiri: Dict[str, int] = {}

    def yaz_kategori_blogu(kat):
        nonlocal row, toplam

        if kat not in kat_satirlar:
            return
        yazilan.add(kat)
        kategori_baslangic_satiri[kat] = row + 1

        # Kategori başlığı - görseldeki gibi düz metin, sadece A sütununda, kalın
        cc = ws.cell(row=row, column=1, value=kat)
        cc.font = KAT_FONT
        cc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        for isim in kat_satirlar[kat]:
            key = (kat, isim)
            mc = ws.cell(row=row, column=1, value=isim)
            mc.font = MASRAF_FONT
            mc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            sb(mc)

            col = 2
            degerler = []
            for banka in BANKALAR:
                bd = veri.get(key, {}).get(banka, {"mobil": "", "sube": ""})
                degerler.append((banka, "mobil", bd.get("mobil", "")))
                degerler.append((banka, "sube",  bd.get("sube", "")))

            mobil_norm = [_norm_deger(v.split("\n")[0]) for (_, k, v) in degerler if k == "mobil" and v]
            sube_norm  = [_norm_deger(v.split("\n")[0]) for (_, k, v) in degerler if k == "sube"  and v]
            mobil_fark = len(set(mobil_norm)) > 1
            sube_fark  = len(set(sube_norm)) > 1

            for i, (banka, kanal, d) in enumerate(degerler):
                cell_value = d if d else ("N/A" if not d and any(v for (_, _, v) in degerler) else "")
                c = ws.cell(row=row, column=col, value=cell_value)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                sb(c)

                fark = mobil_fark if kanal == "mobil" else sube_fark
                if fark and d:
                    c.fill = YELLOW
                    c.font = Font(size=10, color="FF0000", bold=True)
                else:
                    c.font = Font(size=10, color=BANKA_VERI_RENK[banka],
                                  bold=(kanal == "mobil"))
                col += 1

            ws.row_dimensions[row].height = 30 if any("\n" in v for (_, _, v) in degerler) else 20
            row += 1
            toplam += 1

        row += 1

    for kat in KATEGORI_SIRA:
        yaz_kategori_blogu(kat)

    for kat in list(kat_satirlar.keys()):
        if kat not in yazilan:
            yaz_kategori_blogu(kat)

    # ---- Notlar sütunu (kırmızı italik, görseldeki gibi) ----
    for kat, notlar in KATEGORI_NOTLARI.items():
        baslangic = kategori_baslangic_satiri.get(kat)
        if not baslangic:
            continue
        for i, not_metni in enumerate(notlar):
            nc = ws.cell(row=baslangic + i, column=NOT_KOLONU, value=not_metni)
            nc.font = Font(size=9, color="FF0000", italic=True)
            nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.cell(row=row, column=1,
            value=f"Son güncelleme: {_bugun()}").font = Font(size=9, color="888888")

    wb.save(dosya_yolu)
    print(f"[excel] {dosya_yolu} kaydedildi. {toplam} satır yazıldı.")
